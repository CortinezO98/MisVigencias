import io
from datetime import datetime
from django.http import HttpResponse
from django.template.loader import get_template
from xhtml2pdf import pisa
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
import pandas as pd

def generate_vigencias_pdf(request, vigencias):
    """Genera PDF de vigencias"""
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="vigencias_{datetime.now():%Y%m%d}.pdf"'
    
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=1  # Centrado
    )
    
    # Contenido del PDF
    story = []
    
    # Título
    story.append(Paragraph("Reporte de Vigencias - Mis Vigencias", title_style))
    story.append(Paragraph(f"Generado: {datetime.now():%d/%m/%Y %H:%M}", styles['Normal']))
    story.append(Paragraph(f"Usuario: {request.user.get_full_name() or request.user.username}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Tabla de vigencias
    data = [['Vehículo', 'Tipo', 'Vencimiento', 'Días Rest.', 'Estado']]
    
    for v in vigencias:
        dias = (v.fecha_vencimiento - datetime.now().date()).days
        
        if dias < 0:
            estado = "Vencido"
            color = colors.red
        elif dias <= 7:
            estado = "Por vencer"
            color = colors.orange
        else:
            estado = "Vigente"
            color = colors.green
        
        data.append([
            v.vehicle.alias,
            v.get_tipo_display(),
            v.fecha_vencimiento.strftime('%d/%m/%Y'),
            str(dias),
            estado
        ])
    
    # Crear tabla
    table = Table(data, colWidths=[2*inch, 1.5*inch, 1.5*inch, 1*inch, 1.5*inch])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.black),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
    ]))
    
    story.append(table)
    story.append(Spacer(1, 30))
    
    # Estadísticas
    story.append(Paragraph("Estadísticas", styles['Heading2']))
    
    total = len(vigencias)
    vencidos = sum(1 for v in vigencias if (v.fecha_vencimiento - datetime.now().date()).days < 0)
    proximos = sum(1 for v in vigencias if 0 <= (v.fecha_vencimiento - datetime.now().date()).days <= 7)
    
    stats_data = [
        ['Total Vigencias', str(total)],
        ['Vencidos', str(vencidos)],
        ['Próximos a vencer (≤7 días)', str(proximos)],
        ['Vigentes', str(total - vencidos - proximos)]
    ]
    
    stats_table = Table(stats_data, colWidths=[3*inch, 1*inch])
    stats_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.lightblue),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
    ]))
    
    story.append(stats_table)
    
    # Pie de página
    story.append(Spacer(1, 50))
    story.append(Paragraph("Mis Vigencias - Sistema de gestión vehicular", styles['Normal']))
    story.append(Paragraph("www.misvigencias.co", styles['Normal']))
    
    # Generar PDF
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    
    response.write(pdf)
    return response

def generate_vigencias_excel(request, vigencias):
    """Genera Excel de vigencias"""
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="vigencias_{datetime.now():%Y%m%d}.xlsx"'
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Vigencias"
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Encabezados
    headers = ['Vehículo', 'Placa', 'Tipo', 'Fecha Vencimiento', 
               'Días Restantes', 'Estado', 'Activo', 'Creado']
    
    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = alignment
        cell.border = thin_border
    
    # Datos
    for row_num, vigencia in enumerate(vigencias, 2):
        dias = (vigencia.fecha_vencimiento - datetime.now().date()).days
        
        if dias < 0:
            estado = "Vencido"
        elif dias <= 7:
            estado = "Por vencer"
        else:
            estado = "Vigente"
        
        ws.cell(row=row_num, column=1, value=vigencia.vehicle.alias)
        ws.cell(row=row_num, column=2, value=vigencia.vehicle.plate or "")
        ws.cell(row=row_num, column=3, value=vigencia.get_tipo_display())
        ws.cell(row=row_num, column=4, value=vigencia.fecha_vencimiento)
        ws.cell(row=row_num, column=5, value=dias)
        ws.cell(row=row_num, column=6, value=estado)
        ws.cell(row=row_num, column=7, value="Sí" if vigencia.activo else "No")
        ws.cell(row=row_num, column=8, value=vigencia.created_at.strftime('%d/%m/%Y'))
    
    # Ajustar ancho de columnas
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 30)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Hoja de estadísticas
    ws2 = wb.create_sheet(title="Estadísticas")
    
    total = len(vigencias)
    vencidos = sum(1 for v in vigencias if (v.fecha_vencimiento - datetime.now().date()).days < 0)
    proximos = sum(1 for v in vigencias if 0 <= (v.fecha_vencimiento - datetime.now().date()).days <= 7)
    
    stats = [
        ["Métrica", "Valor"],
        ["Total Vigencias", total],
        ["Vencidos", vencidos],
        ["Próximos a vencer (≤7 días)", proximos],
        ["Vigentes", total - vencidos - proximos]
    ]
    
    for row_num, stat_row in enumerate(stats, 1):
        for col_num, value in enumerate(stat_row, 1):
            ws2.cell(row=row_num, column=col_num, value=value)
            if row_num == 1:
                ws2.cell(row=row_num, column=col_num).font = header_font
    
    # Guardar
    wb.save(response)
    return response

def generate_vehicle_report_pdf(vehicle):
    """Reporte detallado de un vehículo"""
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    
    story = []
    
    # Título
    story.append(Paragraph(f"Reporte de Vehículo: {vehicle.alias}", styles['Heading1']))
    if vehicle.plate:
        story.append(Paragraph(f"Placa: {vehicle.plate}", styles['Normal']))
    story.append(Paragraph(f"Propietario: {vehicle.owner.get_full_name()}", styles['Normal']))
    story.append(Spacer(1, 20))
    
    # Vigencias activas
    vigencias = vehicle.vigencias.filter(activo=True)
    
    if vigencias.exists():
        story.append(Paragraph("Vigencias Activas", styles['Heading2']))
        
        data = [['Tipo', 'Vencimiento', 'Días Rest.', 'Recordatorios']]
        for v in vigencias:
            dias = (v.fecha_vencimiento - datetime.now().date()).days
            reminders = []
            if v.r30: reminders.append("30d")
            if v.r15: reminders.append("15d")
            if v.r7: reminders.append("7d")
            if v.r1: reminders.append("1d")
            
            data.append([
                v.get_tipo_display(),
                v.fecha_vencimiento.strftime('%d/%m/%Y'),
                str(dias),
                ", ".join(reminders)
            ])
        
        table = Table(data, colWidths=[2*inch, 2*inch, 1*inch, 1.5*inch])
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ]))
        story.append(table)
    else:
        story.append(Paragraph("No hay vigencias activas", styles['Normal']))
    
    story.append(Spacer(1, 30))
    
    # Histórico (vigencias inactivas)
    historico = vehicle.vigencias.filter(activo=False)
    if historico.exists():
        story.append(Paragraph("Histórico de Vigencias", styles['Heading2']))
        
        data = [['Tipo', 'Fecha Vencimiento', 'Estado', 'Creado']]
        for v in historico[:10]:  # Últimas 10
            data.append([
                v.get_tipo_display(),
                v.fecha_vencimiento.strftime('%d/%m/%Y'),
                "Renovado" if not v.activo else "Activo",
                v.created_at.strftime('%d/%m/%Y')
            ])
        
        table = Table(data)
        table.setStyle(TableStyle([
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ]))
        story.append(table)
    
    doc.build(story)
    pdf = buffer.getvalue()
    buffer.close()
    
    return pdf